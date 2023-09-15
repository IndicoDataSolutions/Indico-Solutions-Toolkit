import pandas as pd
import openpyxl


def table_to_df(cell_dict):
    table_dict = {}

    for _, value in cell_dict.items():
        rows = value["rows"]
        cols = value["columns"]
        if len(rows) > 1 or len(cols) > 1:
            raise ValueError(
                "Cannot represent a table where cells span multiple rows or columns as a dataframe.  Please consider using an HTML or XLSX representation."
            )
        else:
            row = rows[0]
            col = cols[0]
        text = value["text"]
        table_dict[(row, col)] = text

    max_row = max([key[0] for key in table_dict.keys()]) + 1
    max_col = max([key[1] for key in table_dict.keys()]) + 1

    df = pd.DataFrame(index=range(max_row), columns=range(max_col))

    for (row, col), text in table_dict.items():
        df.at[row, col] = text

    return df


def json_to_html_table(data):
    max_row = max([value["rows"][-1] for _, value in data.items()]) + 1
    max_col = max([value["columns"][-1] for _, value in data.items()]) + 1

    table = [["" for _ in range(max_col)] for _ in range(max_row)]

    for key, value in data.items():
        rows = value["rows"]
        cols = value["columns"]
        text = value["text"]
        rowspan = len(rows)
        colspan = len(cols)
        table[rows[0]][cols[0]] = (text, rowspan, colspan)

    html = "<table border='1'>\n"
    for row in table:
        html += "<tr>\n"
        for cell in row:
            if cell:
                text, rowspan, colspan = cell
                if rowspan > 1 or colspan > 1:
                    html += f"<td rowspan='{rowspan}' colspan='{colspan}'>{text}</td>\n"
                else:
                    html += f"<td>{text}</td>\n"
            else:
                html += "<td></td>\n"
        html += "</tr>\n"
    html += "</table>"

    return html


def json_to_xlsx(data, filename="output.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active

    for key, value in data.items():
        rows = value["rows"]
        cols = value["columns"]
        text = value["text"]

        # Set the value for the top-left cell of the span
        ws.cell(row=rows[0] + 1, column=cols[0] + 1, value=text)

        # Merge cells if they span multiple rows or columns
        if len(rows) > 1 or len(cols) > 1:
            ws.merge_cells(
                start_row=rows[0] + 1,
                start_column=cols[0] + 1,
                end_row=rows[-1] + 1,
                end_column=cols[-1] + 1,
            )

    wb.save(filename)


def get_columns(data):
    columns = {}
    for _, value in data.items():
        text = value["text"]
        for col in value["columns"]:
            if col not in columns:
                columns[col] = []
            columns[col].append(text)
    return columns


def get_rows(data):
    rows = {}
    for _, value in data.items():
        text = value["text"]
        for row in value["rows"]:
            if row not in rows:
                rows[row] = []
            rows[row].append(text)
    return rows


class CellLabelAssociation:
    def __init__(self, label, cells):
        self.label = label
        self.cells = cells

    @classmethod
    def from_dict(cls, data):
        return cls(data["label"], data["cells"])

    def to_dict(self):
        return {"label": self.label, "cells": self.cells}

    def __str__(self):
        cell_indices = list(self.cells.keys())
        return f"Label: {self.label} is in Cells: {', '.join([str(c) for c in cell_indices])}"

    def __repr__(self):
        return self.__str__()

    def __eq__(self, other):
        if isinstance(other, CellLabelAssociation):
            return self.label == other.label and self.cells == other.cells
        return False

    def __hash__(self):
        return hash((self.label, tuple(self.cells.keys())))

    def cells_for_label(self):
        return self.cells

    def label_text(self):
        return self.label["text"]

    def cell_details_by_index(self, index):
        return self.cells.get(index, None)

    def label_type(self):
        return self.label["label"]

    def cells_text(self):
        return " ".join([c["text"] for _, c in self.cells.items()])

    def label_offset(self):
        return {"start": self.label["start"], "end": self.label["end"]}

    def cell_offsets(self):
        return [
            {"start": c["doc_offsets"]["start"], "end": c["doc_offsets"]["end"]}
            for _, c in self.cells.items()
        ]
